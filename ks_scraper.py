#!/usr/bin/env python3
"""
快手视频评论通用抓取器
========================
基于 Playwright + GraphQL API 的通用快手评论抓取工具。

核心思路：
  1. 通过 CDP 连接已登录的浏览器（复用登录态）
  2. GraphQL API 分页拉取根评论
  3. CDP 浏览器 DOM 抓取子回复（点击"展开回复"）
  4. 输出到 Excel，文件名用视频标题（≤10字）
"""

import json
import time
import re
import sys
import os
import socket
import subprocess
import shutil
from datetime import datetime
from typing import Optional
from urllib.parse import urlparse

# Windows GBK 控制台兼容 emoji
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ───────────────────────────────────────
# 常量
# ───────────────────────────────────────

GRAPHQL_URL = "https://www.kuaishou.com/graphql"

# 评论列表 GraphQL 查询
# 注意: headurl 即头像链接
COMMENT_LIST_QUERY = """
query commentListQuery($photoId: String, $pcursor: String) {
  visionCommentList(photoId: $photoId, pcursor: $pcursor) {
    commentCount
    commentCountV2
    pcursor
    pcursorV2
    rootCommentsV2 {
      commentId
      content
      likedCount
      timestamp
      authorName
      headurl
      authorId
      subCommentCount
    }
  }
}
"""

# 视频标题 GraphQL 查询
VIDEO_TITLE_QUERY = """
query videoTitleQuery($photoId: String) {
  visionVideoDetail(photoId: $photoId) {
    photo {
      caption
    }
  }
}
"""

# 模拟浏览器的请求头
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0"
    ),
    "Content-Type": "application/json",
    "Accept": "*/*",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Origin": "https://www.kuaishou.com",
    "Referer": "https://www.kuaishou.com/",
}


# ───────────────────────────────────────
# Edge CDP 自动启动
# ───────────────────────────────────────

EDGE_CDP_PROFILE = os.path.join(os.path.expanduser("~"), ".qclaw", "edge_cdp_profile")


def _find_edge() -> str:
    """查找 Edge 浏览器可执行文件"""
    paths = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
        "/usr/bin/microsoft-edge",
    ]
    for p in paths:
        if os.path.isfile(p):
            return p
    found = shutil.which("msedge") or shutil.which("microsoft-edge")
    if found:
        return found
    raise FileNotFoundError("找不到 Edge 浏览器，请确保已安装 Microsoft Edge")


def _port_open(host: str, port: int, timeout: float = 2.0) -> bool:
    """检查 TCP 端口是否可连接"""
    try:
        s = socket.create_connection((host, port), timeout=timeout)
        s.close()
        return True
    except (ConnectionError, OSError, TimeoutError):
        return False


def _wait_cdp(cdp_url: str, timeout_sec: int = 30) -> bool:
    """阻塞等待 CDP 端口就绪"""
    parsed = urlparse(cdp_url)
    host = parsed.hostname or "127.0.0.1"
    port = parsed.port
    for _ in range(timeout_sec):
        time.sleep(1)
        if _port_open(host, port):
            return True
    return False


def launch_edge_cdp(cdp_url: str = "http://127.0.0.1:28800") -> str:
    """
    启动 Edge 浏览器并开启 CDP（使用独立 profile）。

    - 独立 profile 不影响日常浏览器
    - 首次使用需在打开的窗口中登录快手
    - 登录态持久保留在 ~/.qclaw/edge_cdp_profile 中
    """
    parsed = urlparse(cdp_url)
    host = parsed.hostname or "127.0.0.1"
    port = parsed.port

    if _port_open(host, port):
        return cdp_url  # 已就绪

    edge = _find_edge()
    os.makedirs(EDGE_CDP_PROFILE, exist_ok=True)

    cmd = [
        edge,
        f"--remote-debugging-port={port}",
        f"--user-data-dir={EDGE_CDP_PROFILE}",
        "--no-first-run",
        "--no-default-browser-check",
        "https://www.kuaishou.com",
    ]
    subprocess.Popen(
        cmd,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0,
    )

    if not _wait_cdp(cdp_url, timeout_sec=30):
        raise RuntimeError(
            f"Edge 启动超时（CDP 端口 {port} 未响应）。\n"
            "请关闭已有 Edge 窗口后重试。"
        )

    return cdp_url


# ───────────────────────────────────────
# 工具函数
# ───────────────────────────────────────

def safenum(v):
    """安全转 int（兼容 int / str / None）"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v))
    except (ValueError, TypeError):
        return 0


def safestr(v):
    """安全转 str"""
    if v is None:
        return ""
    return str(v)


def extract_photo_id(url_or_id: str) -> str:
    """
    从快手视频 URL 或直接 ID 中提取 photoId。

    支持格式：
      - 直接 ID: "3xqi7iru65ut3bk"
      - 完整 URL: "https://www.kuaishou.com/short-video/3xqi7iru65ut3bk"
      - 短链: "https://v.kuaishou.com/xxxxx"
      - 旧版 URL: "https://live.kuaishou.com/u/xxx/3xqi7iru65ut3bk"
    """
    url_or_id = url_or_id.strip()

    # 已经是纯 ID（只含字母数字和特殊字符）
    if re.match(r'^[a-zA-Z0-9_\-]+$', url_or_id) and '/' not in url_or_id:
        return url_or_id

    # 从 URL 中提取
    patterns = [
        r'/short-video/([a-zA-Z0-9_\-]+)',
        r'/fw/photo/([a-zA-Z0-9_\-]+)',
        r'photoId=([a-zA-Z0-9_\-]+)',
    ]
    for pat in patterns:
        m = re.search(pat, url_or_id)
        if m:
            return m.group(1)

    raise ValueError(f"无法从输入中提取视频 ID: {url_or_id}")


def timestamp_to_str(ts) -> str:
    """毫秒时间戳 → 可读字符串"""
    ts = safenum(ts)
    if ts == 0:
        return ""
    if ts > 10_000_000_000_000:  # 毫秒过长，可能是微秒
        ts = ts // 1000
    if ts > 10_000_000_000:  # 毫秒
        ts = ts // 1000
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except (OSError, ValueError):
        return str(ts)


def sanitize_filename(title: str, max_len: int = 10) -> str:
    """
    清理视频标题转为安全文件名。

    - 移除 Windows 非法字符: \\ / : * ? \" < > |
    - 去换行、空白 trim
    - 截断到 max_len 个字符
    - 空则返回空串（调用方回退到 photoId）
    """
    if not title:
        return ""
    # 移除非法文件名字符
    title = re.sub(r'[\\/:*?"<>|#]', '', title)
    # 换行 / 制表符 → 空格
    title = re.sub(r'[\r\n\t]+', ' ', title)
    # 多个空格合并
    title = re.sub(r'\s+', ' ', title).strip()
    if len(title) > max_len:
        title = title[:max_len].strip()
    return title


# ───────────────────────────────────────
# 核心类
# ───────────────────────────────────────

class KuaishouCommentScraper:
    """
    快手视频评论抓取器

    参数:
        cdp_url: 浏览器 CDP 地址（默认 http://127.0.0.1:28800）
        cookies: 手动传入的 cookie dict（可选，有 CDP 则自动从浏览器获取）
        cookie_file: cookie 持久化文件路径
        timeout: 请求超时秒数（默认 15）
        page_delay: 翻页间隔秒数（默认 0.5，避免触发风控）
    """

    def __init__(
        self,
        cdp_url: str = "http://127.0.0.1:28800",
        cookies: Optional[dict] = None,
        cookie_file: Optional[str] = None,
        timeout: int = 15,
        page_delay: float = 0.5,
    ):
        self.cdp_url = cdp_url
        self.timeout = timeout
        self.page_delay = page_delay
        self.cookie_file = cookie_file
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self._cookies = cookies
        self._playwright_connected = False
        self._browser = None
        self._context = None
        self._pw = None

    # ── Cookie 持久化 ──────────────────

    def save_cookies(self, path: str = "") -> str:
        """
        从浏览器提取快手 cookie 并保存到文件。
        之后无需 CDP 也能抓取（直到 cookie 过期）。
        """
        path = path or self.cookie_file or "ks_cookies.json"
        cookies = self._get_cookies_from_browser()
        self._cookies = cookies
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cookies, f, ensure_ascii=False, indent=2)
        print(f"  Cookie 已保存到: {path} ({len(cookies)} 个)")
        return path

    def load_cookies(self, path: str = "") -> dict:
        """从文件加载 cookie"""
        path = path or self.cookie_file or "ks_cookies.json"
        if not os.path.exists(path):
            raise FileNotFoundError(f"Cookie 文件不存在: {path}")
        with open(path, "r", encoding="utf-8") as f:
            cookies = json.load(f)
        self._cookies = cookies
        print(f"  已加载 cookie: {path} ({len(cookies)} 个)")
        return cookies

    # ── 浏览器连接 ──────────────────────

    def _ensure_browser(self):
        """确保 Playwright 已连接到浏览器，如未运行则自动启动"""
        if self._playwright_connected:
            return
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            raise ImportError(
                "需要安装 Playwright: pip install playwright && playwright install chromium"
            )

        parsed = urlparse(self.cdp_url)
        host = parsed.hostname or "127.0.0.1"
        port = parsed.port

        self._pw = sync_playwright().start()

        # 先尝试连接已有 CDP，失败则自动启动浏览器
        if not _port_open(host, port):
            print("  浏览器未运行，自动启动 Edge CDP...（首次使用请登录快手）")
            launch_edge_cdp(self.cdp_url)

        self._browser = self._pw.chromium.connect_over_cdp(self.cdp_url)
        contexts = self._browser.contexts
        if not contexts:
            raise RuntimeError("浏览器没有打开的上下文，请先在浏览器中打开快手页面")
        self._context = contexts[0]
        self._playwright_connected = True
        print(f"[✓] 已连接浏览器 (CDP: {self.cdp_url})")

    def _get_cookies_from_browser(self) -> dict:
        """直接从浏览器提取 cookie（需 CDP）"""
        self._ensure_browser()
        all_cookies = self._context.cookies()
        ks_cookies = {}
        for c in all_cookies:
            if "kuaishou.com" in c.get("domain", ""):
                ks_cookies[c["name"]] = c["value"]
        if not ks_cookies:
            print("[!] 警告：浏览器中没有找到快手 cookie，可能无法获取评论")
        return ks_cookies

    def _get_cookies(self) -> dict:
        """
        获取快手域名 cookie。
        优先级: 手动传入 > 文件加载 > 浏览器 CDP 提取
        CDP 提取后自动保存到文件（下次可直接用）
        """
        if self._cookies:
            return self._cookies

        # 尝试从文件加载
        if self.cookie_file and os.path.exists(self.cookie_file):
            try:
                return self.load_cookies(self.cookie_file)
            except Exception:
                pass

        # 回退到浏览器 CDP
        cookies = self._get_cookies_from_browser()
        self._cookies = cookies

        # 自动保存，下次无需 CDP
        if cookies:
            save_path = self.cookie_file or "ks_cookies.json"
            try:
                with open(save_path, "w", encoding="utf-8") as f:
                    json.dump(cookies, f, ensure_ascii=False, indent=2)
                print(f"  Cookie 已自动保存到: {save_path}")
            except Exception:
                pass

        return self._cookies

    # ── 解析分享链接 ───────────────────

    @staticmethod
    def resolve_share_link(url: str, timeout: int = 10) -> str:
        """
        解析快手分享短链接，跟随重定向拿到 photoId。

        支持格式:
          - https://www.kuaishou.com/f/X3dGSlY9A6McJpQ   (分享短链)
          - https://v.kuaishou.com/xxxx                    (短链变体)
          - 普通视频 URL（直接提取 photoId）

        Args:
            url: 快手视频链接
            timeout: HTTP 超时秒数
        Returns:
            photoId 字符串
        Raises:
            ValueError: URL 不含可识别的视频 ID
        """
        url = url.strip()
        # 先看看是不是已经包含 photoId
        try:
            return extract_photo_id(url)
        except ValueError:
            pass

        # 跟随重定向
        try:
            resp = requests.head(
                url,
                allow_redirects=True,
                timeout=timeout,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                },
            )
            final_url = resp.url
        except Exception:
            # HEAD 可能被拒，降级为 GET
            resp = requests.get(
                url,
                allow_redirects=True,
                timeout=timeout,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                },
            )
            final_url = resp.url

        return extract_photo_id(final_url)

    # ── 自动连接 ────────────────────────

    def auto_connect(self) -> bool:
        """
        自动建立连接，无需人工干预。

        策略: 已保存 cookie 文件 → 已有 CDP 端口 → 自动启动 CDP 浏览器
        返回 True 表示已就绪。
        """
        cookie_path = self.cookie_file or "ks_cookies.json"

        # 1. 优先 cookie 文件（零依赖，最快）
        if os.path.exists(cookie_path):
            try:
                self.load_cookies(cookie_path)
                print("  已加载 Cookie，无需浏览器")
                return True
            except Exception:
                print("[!] Cookie 文件损坏，尝试 CDP...")

        # 2. 检查 CDP 端口是否已开放
        parsed = urlparse(self.cdp_url)
        host = parsed.hostname or "127.0.0.1"
        port = parsed.port

        if not _port_open(host, port):
            # 3. 自动启动 CDP 浏览器
            print("  自动启动浏览器...（首次使用请登录快手）")
            launch_edge_cdp(self.cdp_url)

        # 4. 从 CDP 提取 cookie 并自动保存
        try:
            self.save_cookies(cookie_path)
            print("  连接成功，Cookie 已自动保存")
            return True
        except Exception as e:
            raise RuntimeError(
                f"无法获取快手 Cookie。\n"
                f"请确认已在弹出的浏览器中登录快手。\n"
                f"错误: {e}"
            )

    # ── 视频标题 ────────────────────────

    def get_video_title(self, photo_id: str) -> str:
        """
        通过 GraphQL 获取视频说明/标题。

        Args:
            photo_id: 视频 ID
        Returns:
            视频标题字符串，失败返回空字符串
        """
        cookies = self._get_cookies()
        payload = {
            "query": VIDEO_TITLE_QUERY,
            "variables": {"photoId": photo_id},
        }
        try:
            resp = self.session.post(
                GRAPHQL_URL,
                json=payload,
                cookies=cookies,
                timeout=self.timeout,
            )
            data = resp.json()
            photo = data.get("data", {}).get("visionVideoDetail", {}).get("photo", {})
            return photo.get("caption", "") or ""
        except Exception:
            return ""

    # ── 评论抓取 ────────────────────────

    def fetch_comments(
        self,
        video_url_or_id: str,
        max_pages: int = 50,
        verbose: bool = True,
    ) -> list[dict]:
        """
        分页拉取视频的全部根评论。

        参数:
            video_url_or_id: 视频 URL 或 photoId
            max_pages: 最大翻页数（防止死循环）
            verbose: 是否打印进度

        返回:
            list[dict]，每条评论包含:
              - comment_id
              - username
              - avatar
              - content
              - time (可读的日期时间字符串)
              - likes (点赞数)
              - reply_count (子回复数)
              - author_id
        """
        photo_id = extract_photo_id(video_url_or_id)
        cookies = self._get_cookies()

        all_comments = []
        pcursor = ""
        page = 0

        if verbose:
            print(f"\n{'='*60}")
            print(f"  视频 ID: {photo_id}")
            print(f"{'='*60}")

        while page < max_pages:
            page += 1
            if verbose:
                print(f"  第 {page} 页 ...", end=" ")

            # 构造 GraphQL 请求
            payload = {
                "operationName": "commentListQuery",
                "variables": {
                    "photoId": str(photo_id),
                    "pcursor": str(pcursor),
                },
                "query": COMMENT_LIST_QUERY,
            }

            try:
                resp = self.session.post(
                    GRAPHQL_URL,
                    json=payload,
                    cookies=cookies,
                    timeout=self.timeout,
                )
                data = resp.json()
            except Exception as e:
                if verbose:
                    print(f"请求失败: {e}")
                break

            # 解析响应
            comment_list = data.get("data", {}).get("visionCommentList")
            if not comment_list:
                if verbose:
                    print("API 未返回数据（可能需要登录或 cookie 已过期）")
                break

            root_comments = comment_list.get("rootCommentsV2", [])
            new_pcursor = comment_list.get("pcursorV2") or comment_list.get("pcursor", "")
            total_count = comment_list.get("commentCountV2", 0)

            if verbose and page == 1:
                print(f"声称共 {total_count} 条评论")

            if not root_comments:
                if verbose:
                    print("无更多评论")
                break

            # 提取评论数据
            for c in root_comments:
                comment = {
                    "comment_id": safestr(c.get("commentId")),
                    "username": safestr(c.get("authorName")),
                    "avatar": safestr(c.get("headurl")),
                    "content": safestr(c.get("content")),
                    "time": timestamp_to_str(c.get("timestamp")),
                    "likes": safenum(c.get("likedCount")),
                    "reply_count": safenum(c.get("subCommentCount")),
                    "author_id": safestr(c.get("authorId")),
                }
                all_comments.append(comment)

            if verbose:
                print(f"获取 {len(root_comments)} 条 (累计 {len(all_comments)})")

            # 检查是否有下一页
            if not new_pcursor or str(new_pcursor) in ("", "0", "no_more"):
                if verbose:
                    print(f"   全部拉取完成，共 {len(all_comments)} 条")
                break

            pcursor = new_pcursor
            time.sleep(self.page_delay)

        # 按点赞量降序排列
        all_comments.sort(key=lambda x: x["likes"], reverse=True)

        if verbose:
            total_likes = sum(c["likes"] for c in all_comments)
            print(f"\n{'='*60}")
            print(f"  总计: {len(all_comments)} 条评论, {total_likes} 赞")
            print(f"{'='*60}")

        return all_comments

    # ── 子评论 DOM 抓取 ─────────────────

    def fetch_sub_comments_dom(
        self,
        photo_id: str,
        verbose: bool = True,
    ) -> dict[str, list[dict]]:
        """
        通过 CDP 浏览器打开视频页，点击"展开回复"按钮，抓取全部子评论。

        需要浏览器已连接（CDP 模式）。如果只有 cookie 文件，此方法不可用。

        Args:
            photo_id: 视频 ID
            verbose: 是否打印进度

        Returns:
            {root_comment_id: [sub_comment_dict, ...]}
            sub_comment_dict 包含: comment_id, username, avatar, content, time, likes
        """
        self._ensure_browser()

        video_url = f"https://www.kuaishou.com/short-video/{photo_id}"
        page = self._context.new_page()

        if verbose:
            print(f"\n  🌐 打开视频页（DOM 子评论抓取）...")

        try:
            # 打开页面
            page.goto(video_url, wait_until="domcontentloaded", timeout=30000)

            # 等待评论区出现
            try:
                page.wait_for_selector(".comment-container, [class*='comment']", timeout=15000)
            except Exception:
                if verbose:
                    print("  [!] 评论区未加载，尝试滚动...")
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(2)

            # 反复滚动加载更多评论
            if verbose:
                print("  📜 滚动加载评论...")
            for i in range(30):
                page.evaluate("window.scrollBy(0, 600)")
                time.sleep(0.4)

            # 用 JS 查找并点击所有"展开回复"按钮
            if verbose:
                print("  🔍 查找「展开回复」按钮...")

            expand_result = page.evaluate("""() => {
                const buttons = [];
                // 查找所有包含"展开"文字的可见元素
                const walker = document.createTreeWalker(
                    document.body,
                    NodeFilter.SHOW_ELEMENT,
                    null
                );
                while (walker.nextNode()) {
                    const el = walker.currentNode;
                    if (el.children.length === 0 && el.textContent && el.textContent.includes('展开') && el.textContent.length < 20) {
                        buttons.push(el);
                    }
                }
                // 点击它们（从下到上避免 DOM 变化问题）
                let clicked = 0;
                for (let i = buttons.length - 1; i >= 0; i--) {
                    try {
                        buttons[i].click();
                        clicked++;
                    } catch(e) {}
                }
                return {total: buttons.length, clicked};
            }""")

            if verbose:
                print(f"    找到 {expand_result['total']} 个，点击了 {expand_result['clicked']} 个")

            # 等待子评论加载
            time.sleep(3)

            # 再次滚动确保子评论可见
            page.evaluate("window.scrollBy(0, 300)")
            time.sleep(0.5)
            page.evaluate("window.scrollBy(0, -300)")
            time.sleep(0.5)

            # 用 JS 提取子评论
            if verbose:
                print("  📝 提取子评论数据...")

            sub_data = page.evaluate("""() => {
                const result = {};
                // 找所有子评论 — 快手子评论通常有 subCommentId 作为 data 属性
                const subElements = document.querySelectorAll('[data-sub-comment-id], [class*="subComment"], [class*="sub-comment"]');
                subElements.forEach(el => {
                    const parentCommentEl = el.closest('[data-comment-id]');
                    const rootId = parentCommentEl ? parentCommentEl.getAttribute('data-comment-id') : 'unknown';
                    if (!result[rootId]) result[rootId] = [];

                    const cid = el.getAttribute('data-sub-comment-id') || '';
                    const author = el.querySelector('[class*="name"], [class*="author"], span')?.textContent?.trim() || '';
                    const content = el.querySelector('[class*="content"], [class*="text"], p')?.textContent?.trim() || el.textContent?.trim() || '';
                    const avatar = el.querySelector('img')?.src || '';
                    const timeEl = el.querySelector('[class*="time"], [class*="date"], time');
                    const timeText = timeEl ? timeEl.textContent.trim() : '';
                    const likesEl = el.querySelector('[class*="like"], [class*="count"]');
                    const likesText = likesEl ? likesEl.textContent.trim() : '0';

                    result[rootId].push({
                        comment_id: cid,
                        username: author,
                        avatar: avatar,
                        content: content,
                        time: timeText,
                        likes_text: likesText
                    });
                });
                return result;
            }""")

            if verbose:
                total_subs = sum(len(v) for v in sub_data.values())
                print(f"    ✓ 提取到 {total_subs} 条子评论（{len(sub_data)} 个根评论）")

            return sub_data

        finally:
            page.close()

    # ── Excel 导出 ──────────────────────

    @staticmethod
    def export_excel(
        comments: list[dict],
        output_path: str,
        video_title: str = "",
        video_id: str = "",
        sub_comments: Optional[dict[str, list[dict]]] = None,
    ):
        """
        将评论列表导出为 Excel 文件。

        列: 序号 | 用户名 | 头像链接 | 评论内容 | 评论时间 | 点赞量 | 回复数 | 评论ID

        如果提供了 sub_comments，子评论会直接合并到根评论下方（缩进标记）。
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "评论数据"

        # ── 样式定义 ──
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="微软雅黑", size=10)
        cell_align = Alignment(vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        sub_font = Font(name="微软雅黑", size=10, color="555555")
        sub_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # ── 标题行（可选） ──
        if video_title:
            ws.merge_cells("A1:H1")
            title_cell = ws["A1"]
            title_cell.value = f"快手评论 · {video_title}"
            title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30
            start_row = 2
        else:
            start_row = 1

        # ── 表头 ──
        headers = ["序号", "用户名", "头像链接", "评论内容", "评论时间", "点赞量", "回复数", "评论ID"]
        col_widths = [6, 16, 40, 50, 18, 10, 8, 26]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[start_row].height = 22

        # ── 数据行 ──
        seq = 0
        row = start_row
        for c in comments:
            seq += 1
            row += 1
            values = [
                seq,
                c["username"],
                c["avatar"],
                c["content"],
                c["time"],
                c["likes"],
                c.get("reply_count", 0),
                c["comment_id"],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                if col_idx in (1, 5, 6, 7):
                    cell.alignment = center_align
                else:
                    cell.alignment = cell_align

            # 子评论行
            cid = c["comment_id"]
            if sub_comments and cid in sub_comments:
                subs = sub_comments[cid]
                for sub in subs:
                    seq += 1
                    row += 1
                    sub_content = f"↳ {sub.get('content', '')}"
                    sub_values = [
                        seq,
                        sub.get("username", ""),
                        sub.get("avatar", ""),
                        sub_content,
                        sub.get("time", ""),
                        sub.get("likes_text", ""),
                        "",
                        sub.get("comment_id", ""),
                    ]
                    for col_idx, val in enumerate(sub_values, 1):
                        cell = ws.cell(row=row, column=col_idx, value=val)
                        cell.font = sub_font
                        cell.fill = sub_fill
                        cell.border = thin_border
                        if col_idx in (1, 5, 6, 7):
                            cell.alignment = center_align
                        else:
                            cell.alignment = cell_align

        # ── 冻结首行 ──
        ws.freeze_panes = f"A{start_row + 1}"

        # ── 自动筛选 ──
        ws.auto_filter.ref = f"A{start_row}:H{row}"

        # ── 汇总行 ──
        summary_row = row + 2
        ws.merge_cells(f"A{summary_row}:H{summary_row}")
        total_likes = sum(c["likes"] for c in comments)
        sub_total = sum(len(v) for v in (sub_comments or {}).values())
        summary = (
            f"  共 {len(comments)} 条评论"
        )
        if sub_total:
            summary += f" (含 {sub_total} 条回复)"
        summary += (
            f" | 总点赞 {total_likes} | "
            f"导出时间 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        if video_id:
            summary += f" | 视频ID: {video_id}"
        cell = ws.cell(row=summary_row, column=1, value=summary)
        cell.font = Font(name="微软雅黑", size=10, italic=True, color="666666")

        wb.save(output_path)
        print(f"  已保存: {output_path} ({len(comments)} 条评论)")
        return output_path

    # ── 一键抓取 ────────────────────────

    def scrape(
        self,
        video_url_or_id: str,
        output_path: Optional[str] = None,
        include_sub_comments: bool = False,
        **kwargs,
    ) -> tuple[list[dict], str]:
        """
        一步完成：提取 ID → 抓评论 + 子评论 → 导出 Excel（文件名=视频标题≤10字）

        Args:
            video_url_or_id: 视频 URL 或 photoId
            output_path: 输出路径（不传则自动用视频标题生成）
            include_sub_comments: 是否抓取子回复（需要 CDP 浏览器）
            **kwargs: 传给 fetch_comments 的参数

        Returns:
            (comments_list, output_path)
        """
        photo_id = extract_photo_id(video_url_or_id)

        # 获取视频标题用于文件名
        video_title = ""
        try:
            video_title = self.get_video_title(photo_id)
        except Exception:
            pass

        safe_title = sanitize_filename(video_title, max_len=10)

        if output_path is None:
            if safe_title:
                output_path = f"{safe_title}.xlsx"
            else:
                output_path = f"ks_{photo_id}.xlsx"

        comments = self.fetch_comments(video_url_or_id, **kwargs)

        # 子评论抓取（需要 CDP 浏览器）
        sub_comments = None
        if include_sub_comments:
            print("\n  🔽 抓取子评论（DOM 模式）...")
            try:
                sub_comments = self.fetch_sub_comments_dom(photo_id)
            except Exception as e:
                print(f"  [!] 子评论抓取失败: {e}")
                print("  (继续导出根评论)")

        self.export_excel(
            comments, output_path,
            video_title=video_title,
            video_id=photo_id,
            sub_comments=sub_comments,
        )

        total_subs = sum(len(v) for v in (sub_comments or {}).values())
        if total_subs:
            print(f"   其中 {total_subs} 条子回复已合并到 Excel")

        return comments, output_path


# ───────────────────────────────────────
# 命令行入口
# ───────────────────────────────────────

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="快手视频评论通用抓取器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python ks_scraper.py 3xqi7iru65ut3bk
  python ks_scraper.py https://www.kuaishou.com/short-video/3xqi7iru65ut3bk
  python ks_scraper.py 3xqi7iru65ut3bk -o 进球吧_EP01.xlsx --cdp http://127.0.0.1:28800
  python ks_scraper.py 3xqi7iru65ut3bk --sub    # 包含子评论
        """,
    )
    parser.add_argument("video", help="视频 ID 或完整 URL")
    parser.add_argument("-o", "--output", help="输出 Excel 路径（不传则自动用视频标题）")
    parser.add_argument("--cdp", default="http://127.0.0.1:28800", help="浏览器 CDP 地址")
    parser.add_argument("--max-pages", type=int, default=50, help="最大翻页数")
    parser.add_argument("--timeout", type=int, default=15, help="请求超时（秒）")
    parser.add_argument("--delay", type=float, default=0.5, help="翻页间隔（秒）")
    parser.add_argument("--sub", action="store_true", help="抓取子回复（需要 CDP 浏览器）")

    args = parser.parse_args()

    scraper = KuaishouCommentScraper(
        cdp_url=args.cdp,
        timeout=args.timeout,
        page_delay=args.delay,
    )

    comments, path = scraper.scrape(
        args.video,
        output_path=args.output,
        include_sub_comments=args.sub,
        max_pages=args.max_pages,
    )

    # 打印前 5 条预览
    print("\n  评论预览 (前5条):")
    for i, c in enumerate(comments[:5], 1):
        likes = c["likes"]
        content = c["content"][:60].replace("\n", " ")
        print(f"  {i}. [{likes} ] {c['username']}: {content}...")


if __name__ == "__main__":
    main()
